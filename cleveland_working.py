#   Import required packages
import pandas as pd
from bs4 import BeautifulSoup
import numpy as np
from openpyxl.chart import StockChart, Reference, Series
from openpyxl.chart.error_bar import ErrorBars
from openpyxl.chart.axis import DateAxis, ChartLines
from openpyxl.chart.updown_bars import UpDownBars
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
from openpyxl.chart.data_source import NumDataSource, NumData, NumVal, NumRef

#Read the data
data = pd.read_csv(r"C:\Users\aridding\Excel\Cleveland_Data1.csv", header=0)


#Create a Pivot table by mean score
clean_df = data[["Evaluation_Name","Response_ID","Question_ID","Answer","Comments","Clinician_AD_ID"]]
clean_answer = clean_df.Answer.replace(["-2",-2,"-1",-1],"N/A")
clean_df["Answer"] = clean_answer



s = clean_df['Answer']

#  Working only with numeric data
mask = s.isin(['0', '1', '2', '3', '4', '5'])
clean_df = clean_df[mask]
s_clean = pd.to_numeric(clean_df.Answer)
clean_df["Numeric"] = s_clean
clean_df = clean_df.drop(["Answer","Comments"], 1)


clinician_count = clean_df["Clinician_AD_ID"].unique()
clean_pivot = clean_df.pivot_table(index=["Question_ID"],columns=["Clinician_AD_ID"], aggfunc='mean')
count_group = clean_df.groupby(["Clinician_AD_ID"])["Response_ID"].unique()
clean_pivot = clean_pivot.drop("Response_ID", 1)

#count_group = clean_df.groupby("Clinician_AD_ID")["Response_ID"].count()
count_dic = {}

for idx, df in count_group.items():
    count_dic[idx] = len(df)


#Calculate global mean and global standard deviation by question ID 
av_global_score = clean_df[["Question_ID","Numeric"]]
av_global_score_mean = av_global_score.pivot_table(index=["Question_ID"], aggfunc='mean')
av_global_score_std = av_global_score.pivot_table(index=["Question_ID"], aggfunc='std')



#Determine the range of the global answers -  min and max. 
av_group_min = av_global_score.groupby("Question_ID").min()
#av_group_max = av_global_score.groupby("Question_ID").max()



#   Add Global Average to the clinician specific dataframe
clean_pivot["Global Mean"] = av_global_score_mean
clean_pivot["Global STD"] = question_metrics.STD
clean_pivot["Global Min"] = question_metrics.Min
clean_pivot["Global Max"] = question_metrics.Max



#   Set variables for the pivot table we just created
#   Multiply values by 20 as to convert a rating of "5" to percentage-based.
average_global = clean_pivot["Global Mean"]
std_global = clean_pivot["Global STD"]*20
min_global = clean_pivot["Global Min"]
max_global = clean_pivot["Global Max"]


#   Work on parsing the comments out by clinician and separating into their own worksheets
#   Drop unnnecessary columns
###!!!!!!!#####
clean_data = data.copy()
###!!!!!!!#####
clean_data = clean_data[['Evaluation_Name','Question',"Clinician_AD_ID","Comments"]]


#   Create an empty list to append to 
clean_question = []



for i in range(0, len(clean_data["Question"])-1):
    #   For each item in the column "Question" we need to strip
    #   The HTML tags the surround the text.
    #   Here we will use BeautifulSoup4 to get_text and append to clean_questions
    text = clean_data.Question[i]
    soup = BeautifulSoup(text, "html.parser")
    clean_question.append(soup.get_text())

    
#   Keep all columns except the last column
clean_data = clean_data.iloc[:-1,:]

#   Create a new column "Question" with the clean questions we just stripped
#   Create a variable with the unique question list -- quest_uni
clean_data["Question"] = clean_question
quest_uni = clean_data.Question.unique()
mask = np.ones(len(quest_uni),dtype=bool)
mask[[6, 9]] = False
question_result = quest_uni[mask]


#   Not all respondents repsonded. Here we remove the non-responses to
#   Extract only the actual responses and store as "Filtered_comments"
filtered_comments = clean_data[clean_data.Comments.isnull()==False]

#   Create a list of unique clinicians who are part of that Cleveland Set   
#   Store the unique Questions in a variable - ques
clin_list = filtered_comments.Clinician_AD_ID.unique()

QA = filtered_comments[["Question", "Comments","Clinician_AD_ID"]]


lookup_d = {}


#   Create an empty dictionary to append to
#   Keys are the clinician's name
#   Values are the clinicians comments
d = {}

for name in clin_list:
    clin_df = filtered_comments[filtered_comments.Clinician_AD_ID == name]
    clin_comments = clin_df.Comments
    d[name] = list(clin_comments)
    lookup_d[name] = len(clin_comments)


#   Convert Dictionary to list to make working with easier
ky = []
vl = []
for k, v in d.items():
    ky.append(k)
    vl.append(v)


#   Create a DataFrame with the newly created list
#   Set the index of the Dataframe as the Clinician name
df = pd.DataFrame(vl)
df.index = ky


wb = Workbook()
s1 = wb.get_sheet_by_name("Sheet")
wb.remove_sheet(s1)

summary_of_scores = []
for i in range(0, len(clean_pivot.columns)-4):
    temp_name = clean_pivot.iloc[:,i].name[1]
    temp_scores = clean_pivot.iloc[:,i]*20
    temp_scores = round(temp_scores,2)



#   Create the stacked bar chart by clinician
for i in range(0, len(clean_pivot.columns)-4):
    #   Store clinician's ID into temp_name
    #   Store the Average Score (out of 100) into temp_scores
    #   Store the Global Score (out of 100) into temp_glob_scores
    #   Store the Min/Max/std into respective variables
    temp_name = clean_pivot.iloc[:,i].name[1]
    temp_scores = clean_pivot.iloc[:,i]*20
    temp_scores = round(temp_scores,2)
    temp_glob_scores = average_global*20
    temp_glob_scores = round(temp_glob_scores,2)    
    temp_min = min_global*20
    temp_max = max_global*20
    temp_glob_std = std_global
    
     
    
    #   Create DataFrame, transpose and inset a numbered index
    temp_df = pd.DataFrame([temp_max, temp_min,temp_scores, temp_glob_scores])
    temp_df = temp_df.T


    temp_df.insert(0, "Questions", question_result)

    temp_df.index = range(1, 1+ len(temp_df))
    
    #   Remove irrelevant questions, rename columns and create a deep copy
    temp_df = temp_df.drop(temp_df.index[[6, 9]])
    temp_df.columns = ["Questions","Max","Min","Clinician Average","Average Across Clinicians"]
    temp_df["Clinician Average"].fillna('Missing',inplace=True)

    
    temp_quest_list = temp_df.copy()
    
    #   Create clinician sheet, BarChart obect, style and titles
    ws1 = wb.create_sheet(str(temp_name))
    ws1.column_dimensions["B"].width = 100     # pass an integer
    ws1.page_setup.orientation = ws1.ORIENTATION_LANDSCAPE

    ws1.page_setup.fitToHeight = 0
    ws1.page_setup.fitToWidth = 1
    chart1 = StockChart()
    chart1.type = "col"
    chart1.style = 2
    chart1.width = 20
    chart1.title = str(temp_name) + " Set 5"
    chart1.y_axis.title = 'Score'
    chart1.x_axis.title = 'Question ID'

 
    
    #   Set the reference of the bar charts, set categories and shape
    temp_data = Reference(ws1, min_col=3, min_row=34, max_row=50, max_col=5)
    cats = Reference(ws1, min_col=1, min_row=35, max_row=50)
    chart1.add_data(temp_data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.hiLowLines = ChartLines()
    try:
        ws1["D2"] = int(lookup_d[temp_name])
    except:
        ws1["D2"] = 0
        
    #   Add the chart to cell "A1", below chart add question list
    ws1.add_chart(chart1, "A1")
    ws1["C1"] = "Count of Comment Responses"
    ws1["D1"] = "Count of Score Responses"
    ws1["C2"] = count_dic[temp_name]
    for r in dataframe_to_rows(temp_quest_list, index=True, header=False):
        ws1["A15"] = "Question list"
        ws1["A15"].font = Font(bold=True)
        ws1.append(list(r[:2]))
    pts = [NumVal(idx=i) for i in range(len(data) - 1)]
    cache = NumData(pt=pts)
    chart1.series[-1].val.numRef.numCache = cache

    #   Below the question list, Add the response data
    for r in dataframe_to_rows(temp_df, index=True, header=True):
        ws1["A33"] = "Response Data"
        ws1["A33"].font = Font(bold=True)
        ws1.append(r)


    

#   Add the new section, "Free Form Answers" with both the question
#   And the answers associated together
for r in dataframe_to_rows(QA, index=True, header=False):
    temp_ws = wb[r[3]]
    temp_ws["A51"] = "Free Form Answers"
    temp_ws["A51"].font = Font(bold=True)
    temp_q = r[1]
    temp_a = r[2]

    #   determine the last row in column A to append question
    #   Determine the last row in column A to append answer
    temp_ws["A" + str(len(temp_ws["A"])+1)] = temp_q
    temp_ws["A" + str(len(temp_ws["A"]))].font = Font(bold=True)
    temp_ws.merge_cells("A" + str(len(temp_ws["A"])+1) + ":F" + str(len(temp_ws["A"])+1))
    temp_ws["A" + str(len(temp_ws["A"])+1)] = temp_a
    rd = temp_ws.row_dimensions[len(temp_ws["A"])]
    rd.height = 120
    temp_ws["A" + str(len(temp_ws["A"]))].alignment = Alignment(horizontal="justify", vertical='top')
    
wb.save(r"C:\Users\aridding\Downloads\ClevelandResults1_Oct2017.xlsx")
